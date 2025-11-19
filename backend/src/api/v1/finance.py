"""财务后台业务API接口 (T175-T186)

此模块定义财务人员业务操作相关的API端点。

端点分类:
1. Dashboard数据看板 (T175-T178):
   - GET /v1/finance/dashboard - 今日收入概览
   - GET /v1/finance/dashboard/trends - 月度收入趋势
   - GET /v1/finance/top-customers - 消费金额Top客户
   - GET /v1/finance/customers/{operator_id}/details - 客户详细财务信息

2. 退款审核 (T181-T184):
   - GET /v1/finance/refunds - 退款申请列表
   - GET /v1/finance/refunds/{refund_id} - 退款申请详情
   - POST /v1/finance/refunds/{refund_id}/approve - 批准退款
   - POST /v1/finance/refunds/{refund_id}/reject - 拒绝退款

3. 开票审核 (T185-T186):
   - GET /v1/finance/invoices - 开票申请列表
   - POST /v1/finance/invoices/{invoice_id}/approve - 批准开票
   - POST /v1/finance/invoices/{invoice_id}/reject - 拒绝开票

4. 银行转账审核 (T189f-T189h):
   - GET /v1/finance/bank-transfers - 银行转账申请列表
   - POST /v1/finance/bank-transfers/{transfer_id}/approve - 批准转账
   - POST /v1/finance/bank-transfers/{transfer_id}/reject - 拒绝转账

认证方式:
- JWT Token认证 (Authorization: Bearer {token})
- 用户类型要求: finance
"""

from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status, Form, UploadFile, File, Request
from sqlalchemy.ext.asyncio import AsyncSession

from ...api.dependencies import get_db, require_finance
from ...core import BadRequestException, NotFoundException
# Import all finance schemas from the new finance package
from ...schemas.finance import (
    AuditLogListResponse,
    CustomerFinanceDetails,
    DashboardOverview,
    DashboardTrends,
    InvoiceApproveRequest,
    InvoiceListResponse,
    InvoiceApproveResponse,
    RechargeRequest,
    RechargeResponse,
    RefundApproveRequest,
    RefundApproveResponse,
    RefundDetailsResponse,
    RefundListResponse,
    RefundRejectRequest,
    ReportGenerateRequest,
    ReportGenerateResponse,
    ReportListResponse,
    TopCustomersResponse,
    BankTransferListResponse,
    BankTransferListRequest,
    BankTransferItem,
    ApproveBankTransferRequest,
    RejectBankTransferRequest,
    BankTransferDetailResponse,
    BankTransferStatus
)
from ...services.finance_dashboard_service import FinanceDashboardService
from ...services.finance_invoice_service import FinanceInvoiceService
from ...services.finance_refund_service import FinanceRefundService

router = APIRouter(prefix="/finance", tags=["财务后台"])


# ==================== Dashboard数据看板 (T175-T178) ====================


@router.get(
    "/dashboard",
    response_model=DashboardOverview,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="今日收入概览",
    description="""
    获取今日(UTC 00:00-23:59)的财务数据概览。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **响应数据**:
    - today_recharge: 今日充值总额(字符串格式,精确到分)
    - today_consumption: 今日消费总额
    - today_refund: 今日退款总额
    - today_net_income: 今日净收入(充值-退款)
    - total_operators: 总运营商数量(活跃,未删除)
    - active_operators_today: 今日活跃运营商数(有交易记录)

    **业务规则**:
    - 使用UTC时区计算今日范围
    - 净收入 = 充值 - 退款(不扣除消费,因为消费已在余额中)
    - 活跃运营商 = 今日有任意类型交易记录的运营商
    """,
)
async def get_dashboard(
    token: dict = Depends(require_finance), db: AsyncSession = Depends(get_db)
) -> DashboardOverview:
    """获取今日收入概览API (T175)

    Args:
        token: JWT Token payload (包含sub=finance_id, user_type=finance)
        db: 数据库会话

    Returns:
        DashboardOverview: 今日财务概览数据

    Raises:
        HTTPException 401: 未认证或Token无效
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        dashboard_service = FinanceDashboardService(db)
        overview = await dashboard_service.get_dashboard_overview()
        return overview

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取Dashboard数据失败: {str(e)}",
            },
        )


@router.get(
    "/dashboard/trends",
    response_model=DashboardTrends,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误(月份格式不正确)"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="月度收入趋势",
    description="""
    获取指定月份的每日收入趋势数据,用于绘制趋势图表。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - month: 月份(YYYY-MM格式,可选,默认当前月)

    **响应数据**:
    - month: 查询的月份(YYYY-MM格式)
    - chart_data: 每日数据数组,包含该月所有天(即使无数据也显示0)
      - date: 日期(YYYY-MM-DD)
      - recharge: 当日充值
      - consumption: 当日消费
      - refund: 当日退款
      - net_income: 当日净收入(充值-退款)
    - summary: 月度汇总
      - total_recharge: 月度总充值
      - total_consumption: 月度总消费
      - total_refund: 月度总退款
      - total_net_income: 月度总净收入

    **业务规则**:
    - 返回完整月份数据(1-28/29/30/31日)
    - 无交易的日期显示为"0.00"
    - 使用UTC时区
    """,
)
async def get_dashboard_trends(
    month: Optional[str] = Query(None, description="月份(YYYY-MM格式,默认当前月)"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> DashboardTrends:
    """获取月度收入趋势API (T176)

    Args:
        month: 查询月份(YYYY-MM格式)
        token: JWT Token payload
        db: 数据库会话

    Returns:
        DashboardTrends: 月度趋势数据

    Raises:
        HTTPException 400: 月份格式错误
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        dashboard_service = FinanceDashboardService(db)
        trends = await dashboard_service.get_dashboard_trends(month)
        return trends

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_MONTH", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取趋势数据失败: {str(e)}",
            },
        )


@router.get(
    "/top-customers",
    response_model=TopCustomersResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误(limit超出范围)"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="消费金额Top客户",
    description="""
    获取消费金额排名靠前的客户列表。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - limit: 返回Top N客户(1-100,默认10)

    **响应数据**:
    - customers: Top客户列表(按消费金额降序)
      - rank: 排名(1-N)
      - operator_id: 运营商ID
      - operator_name: 运营商名称
      - category: 客户分类(trial/normal/vip)
      - total_consumption: 累计消费金额
      - consumption_percentage: 占总消费百分比(保留2位小数)
      - total_sessions: 累计游戏场次
    - total_consumption: 所有客户总消费金额

    **业务规则**:
    - 仅统计消费类型交易(consumption)
    - 按消费金额降序排列
    - 百分比 = (客户消费 / 总消费) * 100
    """,
)
async def get_top_customers(
    limit: int = Query(10, ge=1, le=100, description="返回Top N客户"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> TopCustomersResponse:
    """获取消费金额Top客户API (T177)

    Args:
        limit: 返回Top N客户
        token: JWT Token payload
        db: 数据库会话

    Returns:
        TopCustomersResponse: Top客户列表

    Raises:
        HTTPException 400: limit超出范围
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        dashboard_service = FinanceDashboardService(db)
        top_customers = await dashboard_service.get_top_customers(limit=limit)
        return top_customers

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_LIMIT", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取Top客户失败: {str(e)}",
            },
        )


@router.get(
    "/customers/{operator_id}/details",
    response_model=CustomerFinanceDetails,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "运营商ID格式错误"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "运营商不存在"},
    },
    summary="客户详细财务信息",
    description="""
    获取指定客户的详细财务数据。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - operator_id: 运营商ID(UUID格式)

    **响应数据**:
    - operator_id: 运营商ID
    - operator_name: 运营商名称
    - category: 客户分类
    - current_balance: 当前余额
    - total_recharged: 累计充值金额
    - total_consumed: 累计消费金额
    - total_refunded: 累计退款金额
    - total_sessions: 累计游戏场次
    - first_transaction_at: 首次交易时间
    - last_transaction_at: 最近交易时间

    **业务规则**:
    - 统计该运营商所有历史交易记录
    - 用于审核退款时评估客户价值
    """,
)
async def get_customer_finance_details(
    operator_id: str,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> CustomerFinanceDetails:
    """获取客户详细财务信息API (T178)

    Args:
        operator_id: 运营商ID
        token: JWT Token payload
        db: 数据库会话

    Returns:
        CustomerFinanceDetails: 客户财务详情

    Raises:
        HTTPException 400: ID格式错误
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 运营商不存在
        HTTPException 500: 服务器内部错误
    """
    try:
        dashboard_service = FinanceDashboardService(db)
        details = await dashboard_service.get_customer_finance_details(operator_id)
        return details

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_OPERATOR_ID", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "OPERATOR_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取客户财务详情失败: {str(e)}",
            },
        )


# ==================== 退款审核 (T181-T184) ====================


@router.get(
    "/refunds",
    response_model=RefundListResponse,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="退款申请列表",
    description="""
    获取退款申请列表,支持分页和状态筛选。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - status: 状态筛选(pending/approved/rejected/all,默认all)
    - page: 页码(从1开始,默认1)
    - page_size: 每页条数(1-100,默认20)

    **响应数据**:
    - page: 当前页码
    - page_size: 每页条数
    - total: 符合条件的总记录数
    - items: 退款申请列表(按创建时间倒序)
      - refund_id: 退款申请ID
      - operator_id: 运营商ID
      - operator_name: 运营商名称
      - operator_category: 客户分类
      - requested_amount: 申请退款金额
      - current_balance: 当前余额
      - actual_refund_amount: 实际退款金额(已审核时)
      - status: 状态(pending/approved/rejected)
      - reason: 退款原因
      - reject_reason: 拒绝原因(已拒绝时)
      - reviewed_by: 审核人ID
      - reviewed_at: 审核时间
      - created_at: 申请创建时间

    **业务规则**:
    - 按创建时间倒序(最新申请在前)
    - 实际退款金额可能小于申请金额(以当前余额为准)
    """,
)
async def get_refunds(
    status: Optional[str] = Query("all", description="状态筛选(pending/approved/rejected/all)"),
    page: int = Query(1, ge=1, description="页码(从1开始)"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> RefundListResponse:
    """获取退款申请列表API (T181)

    Args:
        status: 状态筛选
        page: 页码
        page_size: 每页条数
        token: JWT Token payload
        db: 数据库会话

    Returns:
        RefundListResponse: 分页的退款申请列表

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        refund_service = FinanceRefundService(db)
        refunds = await refund_service.get_refunds(
            status=status, page=page, page_size=page_size
        )
        return refunds

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取退款列表失败: {str(e)}",
            },
        )


@router.get(
    "/refunds/{refund_id}",
    response_model=RefundDetailsResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "退款ID格式错误"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "退款申请不存在"},
    },
    summary="退款申请详情",
    description="""
    获取退款申请的详细信息,包含运营商财务概况。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - refund_id: 退款申请ID(UUID格式)

    **响应数据**:
    - 退款申请基本信息(同列表接口)
    - operator_finance: 运营商财务概况
      - operator_id: 运营商ID
      - operator_name: 运营商名称
      - category: 客户分类
      - current_balance: 当前余额
      - total_recharged: 累计充值
      - total_consumed: 累计消费
      - total_refunded: 累计退款
      - total_sessions: 累计场次
      - first_transaction_at: 首次交易时间
      - last_transaction_at: 最近交易时间

    **业务规则**:
    - 用于审核前评估客户价值和退款合理性
    - 财务数据统计包含该申请之前的所有历史记录
    """,
)
async def get_refund_details(
    refund_id: str,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> RefundDetailsResponse:
    """获取退款申请详情API (T182)

    Args:
        refund_id: 退款申请ID
        token: JWT Token payload
        db: 数据库会话

    Returns:
        RefundDetailsResponse: 退款申请详情

    Raises:
        HTTPException 400: ID格式错误
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 退款申请不存在
        HTTPException 500: 服务器内部错误
    """
    try:
        refund_service = FinanceRefundService(db)
        details = await refund_service.get_refund_details(refund_id)
        return details

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_REFUND_ID", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "REFUND_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取退款详情失败: {str(e)}",
            },
        )


@router.post(
    "/refunds/{refund_id}/approve",
    response_model=RefundApproveResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误或退款申请状态不允许审核"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "退款申请不存在"},
    },
    summary="批准退款申请",
    description="""
    批准退款申请,扣除运营商余额并创建退款交易记录。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - refund_id: 退款申请ID(UUID格式)

    **请求体**:
    - note: 审核备注(可选,最多200字符)

    **响应数据**:
    - refund_id: 退款申请ID
    - requested_amount: 申请退款金额
    - actual_refund_amount: 实际退款金额(=当前余额)
    - balance_after: 退款后余额(=0.00)

    **业务规则**:
    - 仅能审核pending状态的申请
    - 实际退款金额 = 当前余额(全部退款)
    - 退款后余额归零
    - 创建退款类型交易记录
    - 记录审核人和审核时间
    - 若当前余额为0则无法批准
    """,
)
async def approve_refund(
    refund_id: str,
    request: RefundApproveRequest,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> RefundApproveResponse:
    """批准退款申请API (T183)

    Args:
        refund_id: 退款申请ID
        request: 审核请求(包含可选的note)
        token: JWT Token payload
        db: 数据库会话

    Returns:
        RefundApproveResponse: 审核结果

    Raises:
        HTTPException 400: ID格式错误或状态不允许审核或余额不足
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 退款申请不存在
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        refund_service = FinanceRefundService(db)
        result = await refund_service.approve_refund(
            refund_id=refund_id, finance_id=finance_id, note=request.note
        )
        return result

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "APPROVAL_FAILED", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "REFUND_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"批准退款失败: {str(e)}",
            },
        )


@router.post(
    "/refunds/{refund_id}/reject",
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误或退款申请状态不允许审核"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "退款申请不存在"},
    },
    summary="拒绝退款申请",
    description="""
    拒绝退款申请,不修改运营商余额。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - refund_id: 退款申请ID(UUID格式)

    **请求体**:
    - reason: 拒绝原因(必填,10-200字符)

    **响应**:
    - 无响应体(HTTP 200)

    **业务规则**:
    - 仅能审核pending状态的申请
    - 必须提供拒绝原因
    - 不修改运营商余额
    - 不创建交易记录
    - 记录审核人、审核时间和拒绝原因
    """,
)
async def reject_refund(
    refund_id: str,
    request: RefundRejectRequest,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """拒绝退款申请API (T184)

    Args:
        refund_id: 退款申请ID
        request: 拒绝请求(包含reason)
        token: JWT Token payload
        db: 数据库会话

    Returns:
        dict: 空dict(HTTP 200)

    Raises:
        HTTPException 400: ID格式错误或状态不允许审核或reason格式不正确
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 退款申请不存在
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        refund_service = FinanceRefundService(db)
        await refund_service.reject_refund(
            refund_id=refund_id, finance_id=finance_id, reason=request.reason
        )
        return {}

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "REJECTION_FAILED", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "REFUND_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"拒绝退款失败: {str(e)}",
            },
        )


# ==================== 开票审核 (T185-T186) ====================


@router.get(
    "/invoices",
    response_model=InvoiceListResponse,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="开票申请列表",
    description="""
    获取开票申请列表,支持分页和状态筛选。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - status: 状态筛选(pending/approved/rejected/all,默认all)
    - page: 页码(从1开始,默认1)
    - page_size: 每页条数(1-100,默认20)

    **响应数据**:
    - page: 当前页码
    - page_size: 每页条数
    - total: 符合条件的总记录数
    - items: 开票申请列表(按申请时间倒序)
      - invoice_id: 开票申请ID
      - operator_id: 运营商ID
      - operator_name: 运营商名称
      - operator_category: 客户分类
      - amount: 开票金额
      - invoice_title: 发票抬头
      - tax_id: 纳税人识别号
      - email: 接收邮箱
      - status: 状态(pending/approved/rejected)
      - pdf_url: 发票PDF文件URL(已批准时)
      - reviewed_by: 审核人ID
      - reviewed_at: 审核时间
      - created_at: 申请创建时间

    **业务规则**:
    - 按申请时间倒序(最新申请在前)
    - PDF文件在批准后由后台异步生成
    """,
)
async def get_invoices(
    status: Optional[str] = Query("all", description="状态筛选(pending/approved/rejected/all)"),
    page: int = Query(1, ge=1, description="页码(从1开始)"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> InvoiceListResponse:
    """获取开票申请列表API (T185)

    Args:
        status: 状态筛选
        page: 页码
        page_size: 每页条数
        token: JWT Token payload
        db: 数据库会话

    Returns:
        InvoiceListResponse: 分页的开票申请列表

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        invoice_service = FinanceInvoiceService(db)
        invoices = await invoice_service.get_invoices(
            status=status, page=page, page_size=page_size
        )
        return invoices

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取开票列表失败: {str(e)}",
            },
        )


@router.post(
    "/invoices/{invoice_id}/approve",
    response_model=InvoiceApproveResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误或开票申请状态不允许审核"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "开票申请不存在"},
    },
    summary="批准开票申请",
    description="""
    批准开票申请,生成发票编号和PDF文件URL。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - invoice_id: 开票申请ID(UUID格式)

    **请求体**:
    - note: 审核备注(可选,最多200字符)

    **响应数据**:
    - invoice_id: 开票申请ID
    - pdf_url: 发票PDF文件URL

    **业务规则**:
    - 仅能审核pending状态的申请
    - 生成发票编号: INV-YYYYMMDD-XXXXX
    - 触发异步PDF生成任务(当前返回占位URL)
    - 记录审核人、审核时间和开票时间
    - PDF生成完成后会更新invoice_file_url
    """,
)
async def approve_invoice(
    invoice_id: str,
    request: InvoiceApproveRequest,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> InvoiceApproveResponse:
    """批准开票申请API (T186)

    Args:
        invoice_id: 开票申请ID
        request: 审核请求(包含可选的note)
        token: JWT Token payload
        db: 数据库会话

    Returns:
        InvoiceApproveResponse: 审核结果(包含PDF URL)

    Raises:
        HTTPException 400: ID格式错误或状态不允许审核
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 开票申请不存在
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        invoice_service = FinanceInvoiceService(db)
        result = await invoice_service.approve_invoice(
            invoice_id=invoice_id, finance_id=finance_id, note=request.note
        )
        return result

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "APPROVAL_FAILED", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "INVOICE_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"批准开票失败: {str(e)}",
            },
        )


@router.post(
    "/invoices/{invoice_id}/reject",
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误或开票申请状态不允许审核"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "开票申请不存在"},
    },
    summary="拒绝开票申请",
    description="""
    拒绝开票申请,不生成发票PDF。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - invoice_id: 开票申请ID(UUID格式)

    **请求体**:
    - reason: 拒绝原因(必填,10-200字符)

    **响应**:
    - 无响应体(HTTP 200)

    **业务规则**:
    - 仅能审核pending状态的申请
    - 必须提供拒绝原因
    - 不生成发票编号和PDF
    - 记录审核人、审核时间和拒绝原因
    """,
)
async def reject_invoice(
    invoice_id: str,
    request: RefundRejectRequest,  # Reuse RefundRejectRequest as it has the same structure
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """拒绝开票申请API (T186续)

    Args:
        invoice_id: 开票申请ID
        request: 拒绝请求(包含reason)
        token: JWT Token payload
        db: 数据库会话

    Returns:
        dict: 空dict(HTTP 200)

    Raises:
        HTTPException 400: ID格式错误或状态不允许审核或reason格式不正确
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 开票申请不存在
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        invoice_service = FinanceInvoiceService(db)
        await invoice_service.reject_invoice(
            invoice_id=invoice_id, finance_id=finance_id, reason=request.reason
        )
        return {}

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "REJECTION_FAILED", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "INVOICE_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"拒绝开票失败: {str(e)}",
            },
        )


# ==================== 审计日志 (T167) ====================


@router.get(
    "/audit-logs",
    response_model=AuditLogListResponse,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="获取审计日志列表",
    description="""
    获取财务操作审计日志列表,支持分页和筛选。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - operation_type: 操作类型筛选(可选)
    - page: 页码(从1开始,默认1)
    - page_size: 每页条数(1-100,默认20)

    **响应数据**:
    - page: 当前页码
    - page_size: 每页条数
    - total: 符合条件的总记录数
    - items: 审计日志列表(按时间倒序)

    **业务规则**:
    - 按操作时间倒序(最新记录在前)
    - 记录所有财务人员的关键操作
    """,
)
async def get_audit_logs(
    operation_type: Optional[str] = Query(None, description="操作类型筛选"),
    page: int = Query(1, ge=1, description="页码(从1开始)"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> AuditLogListResponse:
    """获取审计日志列表API (T167)

    Args:
        operation_type: 操作类型筛选
        page: 页码
        page_size: 每页条数
        token: JWT Token payload
        db: 数据库会话

    Returns:
        AuditLogListResponse: 分页的审计日志列表

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        from sqlalchemy import select, func
        from ...models.finance import FinanceOperationLog, FinanceAccount
        from sqlalchemy.orm import selectinload

        # Build query
        query = select(FinanceOperationLog).options(
            selectinload(FinanceOperationLog.finance_account)
        )

        # Apply filters
        if operation_type:
            query = query.where(FinanceOperationLog.operation_type == operation_type)

        # Count total
        count_query = select(func.count()).select_from(FinanceOperationLog)
        if operation_type:
            count_query = count_query.where(FinanceOperationLog.operation_type == operation_type)

        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination and ordering
        query = query.order_by(FinanceOperationLog.created_at.desc())
        query = query.offset((page - 1) * page_size).limit(page_size)

        # Execute query
        result = await db.execute(query)
        logs = result.scalars().all()

        # Build response
        from ...schemas.finance import AuditLogItem
        items = []
        for log in logs:
            items.append(AuditLogItem(
                log_id=f"log_{log.id}",
                finance_id=str(log.finance_account_id),
                finance_name=log.finance_account.full_name if log.finance_account else "Unknown",
                operation_type=log.operation_type,
                target_resource_id=str(log.target_resource_id) if log.target_resource_id else "",
                operation_details=log.operation_details or {},
                ip_address=log.ip_address,
                user_agent=log.user_agent,
                created_at=log.created_at,
            ))

        return AuditLogListResponse(
            page=page,
            page_size=page_size,
            total=total,
            items=items,
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取审计日志失败: {str(e)}",
            },
        )


# ==================== 财务报表 (T166) ====================


@router.post(
    "/reports/generate",
    response_model=ReportGenerateResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="生成财务报表",
    description="""
    生成财务报表(异步任务)。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **请求体**:
    - report_type: 报表类型(daily/weekly/monthly/custom)
    - start_date: 开始日期
    - end_date: 结束日期
    - format: 报表格式(pdf/excel)
    - include_sections: 包含的报表章节(可选)

    **响应数据**:
    - report_id: 报表ID
    - status: 生成状态(generating/completed/failed)
    - estimated_time: 预计完成时间(秒)

    **业务规则**:
    - 报表生成为异步任务
    - 可通过报表ID查询生成状态
    """,
)
async def generate_report(
    request: ReportGenerateRequest,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> ReportGenerateResponse:
    """生成财务报表API (T166)

    Args:
        request: 报表生成请求
        token: JWT Token payload
        db: 数据库会话

    Returns:
        ReportGenerateResponse: 报表生成响应

    Raises:
        HTTPException 400: 请求参数错误
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        from datetime import datetime
        from ...services.finance_report_service import FinanceReportService

        # 将date转换为datetime（如果需要）
        if isinstance(request.start_date, str):
            start_date = datetime.strptime(request.start_date, "%Y-%m-%d")
        else:
            start_date = datetime.combine(request.start_date, datetime.min.time())

        if isinstance(request.end_date, str):
            end_date = datetime.strptime(request.end_date, "%Y-%m-%d")
        else:
            end_date = datetime.combine(request.end_date, datetime.min.time())

        # 设置结束日期为当天的23:59:59
        end_date = end_date.replace(hour=23, minute=59, second=59)

        # 生成报表
        report_service = FinanceReportService(db)
        report = await report_service.generate_report(
            finance_id=finance_id,
            report_type=request.report_type,
            start_date=start_date,
            end_date=end_date,
            export_format=request.format,
        )

        return ReportGenerateResponse(
            report_id=report.report_id,
            status=report.status,
            estimated_time=0,
        )

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_REQUEST", "message": str(e)},
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "INVALID_DATE_FORMAT", "message": f"日期格式错误: {str(e)}"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"生成报表失败: {str(e)}",
            },
        )


@router.get(
    "/reports",
    response_model=ReportListResponse,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="获取报表历史列表",
    description="""
    获取历史生成的报表列表,支持分页。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - page: 页码(从1开始,默认1)
    - page_size: 每页条数(1-100,默认20)

    **响应数据**:
    - page: 当前页码
    - page_size: 每页条数
    - total: 符合条件的总记录数
    - items: 报表列表(按创建时间倒序)

    **业务规则**:
    - 按创建时间倒序(最新报表在前)
    - 仅显示当前用户生成的报表
    """,
)
async def get_reports(
    page: int = Query(1, ge=1, description="页码(从1开始)"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> ReportListResponse:
    """获取报表历史列表API (T166)

    Args:
        page: 页码
        page_size: 每页条数
        token: JWT Token payload
        db: 数据库会话

    Returns:
        ReportListResponse: 分页的报表列表

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 500: 服务器内部错误
    """
    try:
        from ...services.finance_report_service import FinanceReportService
        from ...schemas.finance import FinanceReportItem

        report_service = FinanceReportService(db)
        reports, total = await report_service.get_reports_list(page=page, page_size=page_size)

        # 构建响应数据
        items = []
        for report in reports:
            # Debug: print created_at value
            print(f"DEBUG: report_id={report.report_id}, created_at={report.created_at}, type={type(report.created_at)}")

            items.append(FinanceReportItem(
                report_id=report.report_id,
                report_type=report.report_type,
                start_date=report.start_date.date() if hasattr(report.start_date, 'date') else report.start_date,
                end_date=report.end_date.date() if hasattr(report.end_date, 'date') else report.end_date,
                format=report.export_format,
                status=report.status,
                file_size=report.file_size,
                download_url=f"/finance/reports/{report.report_id}/export" if report.status == "completed" else None,
                error_message=report.error_message,
                total_recharge=report.total_recharge,
                total_consumption=report.total_consumption,
                total_refund=report.total_refund,
                net_income=report.net_income,
                created_by=str(report.generated_by) if report.generated_by else "system",
                created_at=report.created_at,
            ))

        return ReportListResponse(
            page=page,
            page_size=page_size,
            total=total,
            items=items,
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"获取报表列表失败: {str(e)}",
            },
        )


@router.get(
    "/reports/{report_id}/export",
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "报表不存在"},
    },
    summary="导出/下载报表文件",
    description="""
    导出指定报表的文件(PDF/Excel)。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **路径参数**:
    - report_id: 报表ID

    **响应**:
    - 文件流(application/pdf 或 application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)

    **业务规则**:
    - 仅允许下载已完成生成的报表
    - 文件名包含报表ID和日期
    """,
)
async def export_report(
    report_id: str,
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
):
    """导出报表文件API (T166)

    Args:
        report_id: 报表ID
        token: JWT Token payload
        db: 数据库会话

    Returns:
        StreamingResponse: 报表文件流

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 报表不存在
        HTTPException 500: 服务器内部错误
    """
    try:
        from ...services.finance_report_service import FinanceReportService
        from fastapi.responses import StreamingResponse
        import io
        import json

        report_service = FinanceReportService(db)
        report = await report_service.get_report_by_id(report_id)

        # 检查报表状态
        if report.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "REPORT_NOT_READY", "message": f"报表尚未生成完成，当前状态: {report.status}"},
            )

        # 根据导出格式生成文件
        if report.export_format == "csv":
            # 生成CSV文件
            csv_data = _generate_csv_report(report)
            output = io.BytesIO(csv_data.encode('utf-8-sig'))  # 使用UTF-8 BOM避免Excel乱码
            media_type = "text/csv"
            filename = f"财务报表_{report.report_id}.csv"

        elif report.export_format == "excel":
            # 生成Excel文件
            excel_data = await _generate_excel_report(report)
            output = io.BytesIO(excel_data)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"财务报表_{report.report_id}.xlsx"

        else:  # pdf
            # 生成PDF文件
            pdf_data = await _generate_pdf_report(report)
            output = io.BytesIO(pdf_data)
            media_type = "application/pdf"
            filename = f"财务报表_{report.report_id}.pdf"

        output.seek(0)

        # URL encode the filename for proper handling of Chinese characters
        from urllib.parse import quote
        encoded_filename = quote(filename)

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "REPORT_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"导出报表失败: {str(e)}",
            },
        )


def _generate_csv_report(report) -> str:
    """生成CSV格式报表"""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    # 写入标题
    writer.writerow(['财务报表'])
    writer.writerow(['报表编号', report.report_id])
    writer.writerow(['报表类型', report.report_type])
    writer.writerow(['报表周期', report.period])
    writer.writerow([])

    # 写入汇总数据
    writer.writerow(['汇总数据'])
    writer.writerow(['指标', '金额'])
    writer.writerow(['总充值', f'{report.total_recharge:.2f}'])
    writer.writerow(['总消费', f'{report.total_consumption:.2f}'])
    writer.writerow(['总退款', f'{report.total_refund:.2f}'])
    writer.writerow(['净收入', f'{report.net_income:.2f}'])
    writer.writerow([])

    # 写入每日数据
    if report.daily_breakdown:
        writer.writerow(['每日明细'])
        writer.writerow(['日期', '充值', '消费', '退款', '净收入'])
        for day in report.daily_breakdown:
            writer.writerow([
                day['date'],
                day['recharge'],
                day['consumption'],
                day['refund'],
                day['net_income']
            ])
        writer.writerow([])

    # 写入Top客户
    if report.top_customers:
        writer.writerow(['Top消费客户'])
        writer.writerow(['排名', '运营商', '消费金额', '交易笔数'])
        for customer in report.top_customers:
            writer.writerow([
                customer['rank'],
                customer['operator_name'],
                customer['total_consumption'],
                customer['transaction_count']
            ])

    return output.getvalue()


async def _generate_excel_report(report) -> bytes:
    """生成Excel格式报表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "财务报表"

        # 标题样式
        title_font = Font(size=14, bold=True)
        header_font = Font(size=12, bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 写入标题
        ws['A1'] = '财务报表'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        ws[f'A{row}'] = '报表编号:'
        ws[f'B{row}'] = report.report_id
        row += 1
        ws[f'A{row}'] = '报表类型:'
        ws[f'B{row}'] = report.report_type
        row += 1
        ws[f'A{row}'] = '报表周期:'
        ws[f'B{row}'] = report.period
        row += 2

        # 汇总数据
        ws[f'A{row}'] = '汇总数据'
        ws[f'A{row}'].font = header_font
        row += 1

        ws[f'A{row}'] = '指标'
        ws[f'B{row}'] = '金额'
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        row += 1

        ws[f'A{row}'] = '总充值'
        ws[f'B{row}'] = float(report.total_recharge)
        row += 1
        ws[f'A{row}'] = '总消费'
        ws[f'B{row}'] = float(report.total_consumption)
        row += 1
        ws[f'A{row}'] = '总退款'
        ws[f'B{row}'] = float(report.total_refund)
        row += 1
        ws[f'A{row}'] = '净收入'
        ws[f'B{row}'] = float(report.net_income)
        row += 2

        # 每日明细
        if report.daily_breakdown:
            ws[f'A{row}'] = '每日明细'
            ws[f'A{row}'].font = header_font
            row += 1

            headers = ['日期', '充值', '消费', '退款', '净收入']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
            row += 1

            for day in report.daily_breakdown:
                ws.cell(row=row, column=1, value=day['date'])
                ws.cell(row=row, column=2, value=float(day['recharge']))
                ws.cell(row=row, column=3, value=float(day['consumption']))
                ws.cell(row=row, column=4, value=float(day['refund']))
                ws.cell(row=row, column=5, value=float(day['net_income']))
                row += 1
            row += 1

        # Top客户
        if report.top_customers:
            ws[f'A{row}'] = 'Top消费客户'
            ws[f'A{row}'].font = header_font
            row += 1

            headers = ['排名', '运营商', '消费金额', '交易笔数']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
            row += 1

            for customer in report.top_customers:
                ws.cell(row=row, column=1, value=customer['rank'])
                ws.cell(row=row, column=2, value=customer['operator_name'])
                ws.cell(row=row, column=3, value=float(customer['total_consumption']))
                ws.cell(row=row, column=4, value=customer['transaction_count'])
                row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    except ImportError:
        # 如果没有安装openpyxl，返回简单的CSV格式
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail={"error_code": "EXCEL_NOT_SUPPORTED", "message": "Excel导出功能需要安装openpyxl库"},
        )


async def _generate_pdf_report(report) -> bytes:
    """生成PDF格式报表"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import json

    # 注册中文字体
    try:
        # 尝试多个常见的中文字体路径
        font_paths = [
            '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        ]

        font_registered = False
        for font_path in font_paths:
            try:
                import os
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    font_name = 'ChineseFont'
                    font_registered = True
                    break
            except:
                continue

        if not font_registered:
            # 使用reportlab的内置中文字体支持 (STSong-Light)
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            font_name = 'STSong-Light'
    except Exception as e:
        # 最后的备用方案：使用Helvetica（不支持中文，但至少不会崩溃）
        print(f"Warning: Could not register Chinese font: {e}")
        font_name = 'Helvetica'

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)

    # 准备内容
    story = []
    styles = getSampleStyleSheet()

    # 自定义样式
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=14,
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10,
        alignment=TA_LEFT
    )

    # 标题
    story.append(Paragraph('财务报表', title_style))
    story.append(Spacer(1, 0.5*cm))

    # 基本信息
    basic_data = [
        ['报表编号', report.report_id],
        ['报表类型', {'daily': '日报', 'weekly': '周报', 'monthly': '月报', 'custom': '自定义'}.get(report.report_type, report.report_type)],
        ['报表周期', f"{report.start_date.strftime('%Y-%m-%d')} 至 {report.end_date.strftime('%Y-%m-%d')}"],
        ['生成时间', report.generated_at.strftime('%Y-%m-%d %H:%M:%S') if report.generated_at else ''],
    ]

    basic_table = Table(basic_data, colWidths=[4*cm, 12*cm])
    basic_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(basic_table)
    story.append(Spacer(1, 0.8*cm))

    # 汇总数据
    story.append(Paragraph('汇总数据', heading_style))
    summary_data = [
        ['指标', '金额'],
        ['总充值', f'¥{float(report.total_recharge):,.2f}'],
        ['总消费', f'¥{float(report.total_consumption):,.2f}'],
        ['总退款', f'¥{float(report.total_refund):,.2f}'],
        ['净收入', f'¥{float(report.net_income):,.2f}'],
    ]

    summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.8*cm))

    # 每日明细
    if report.daily_breakdown:
        story.append(Paragraph('每日明细', heading_style))
        daily_data = [['日期', '充值', '消费', '退款', '净收入']]

        breakdown = json.loads(report.daily_breakdown) if isinstance(report.daily_breakdown, str) else report.daily_breakdown
        for day in breakdown[:10]:  # 最多显示10天
            daily_data.append([
                day.get('date', ''),
                f"¥{float(day.get('recharge', 0)):,.2f}",
                f"¥{float(day.get('consumption', 0)):,.2f}",
                f"¥{float(day.get('refund', 0)):,.2f}",
                f"¥{float(day.get('net_income', 0)):,.2f}",
            ])

        daily_table = Table(daily_data, colWidths=[3.2*cm, 3.2*cm, 3.2*cm, 3.2*cm, 3.2*cm])
        daily_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(daily_table)
        story.append(Spacer(1, 0.8*cm))

    # Top消费客户
    if report.top_customers:
        story.append(Paragraph('Top消费客户', heading_style))
        top_data = [['排名', '运营商名称', '消费金额']]

        customers = json.loads(report.top_customers) if isinstance(report.top_customers, str) else report.top_customers
        for idx, customer in enumerate(customers[:10], 1):  # 最多显示10个
            top_data.append([
                str(idx),
                customer.get('name', ''),
                f"¥{float(customer.get('consumption', 0)):,.2f}",
            ])

        top_table = Table(top_data, colWidths=[2*cm, 8*cm, 6*cm])
        top_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(top_table)

    # 生成PDF
    doc.build(story)
    output.seek(0)
    return output.read()


# ==================== 运营商列表 ====================


@router.get(
    "/operators",
    response_model=dict,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="获取运营商列表",
    description="""
    获取运营商列表（用于手动充值时选择运营商）。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **查询参数**:
    - search: 搜索关键词（可选），支持用户名、姓名、邮箱、电话
    - status: 状态筛选（可选），active/inactive/locked
    - page: 页码，默认1
    - page_size: 每页条数，默认100（最大1000）

    **响应数据**:
    - items: 运营商列表
      - id: 运营商ID
      - username: 用户名
      - full_name: 真实姓名/公司名称
      - email: 邮箱
      - phone: 电话
      - balance: 当前余额
      - customer_tier: 客户分类
      - is_active: 是否激活
      - is_locked: 是否锁定
      - last_login_at: 最后登录时间
    - total: 总数量
    - page: 当前页码
    - page_size: 每页条数

    **业务规则**:
    - 只返回未删除的运营商
    - 支持关键词搜索和状态筛选
    - 默认按创建时间倒序排列
    """,
)
async def get_operators(
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = Query(None, description="搜索关键词"),
    status_filter: Optional[str] = Query(
        None,
        alias="status",
        description="状态筛选: active/inactive/locked"
    ),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(100, ge=1, le=1000, description="每页条数"),
) -> dict:
    """获取运营商列表API

    Args:
        token: JWT Token payload
        db: 数据库会话
        search: 搜索关键词
        status_filter: 状态筛选
        page: 页码
        page_size: 每页条数

    Returns:
        dict: 运营商列表（分页）

    Raises:
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
    """
    from sqlalchemy import select, or_, func
    from ...models.operator import OperatorAccount

    # 🚀 性能优化: 构建WHERE条件列表,同时用于COUNT和数据查询
    # 原方案: 使用subquery包装完整查询再COUNT,需要执行两次复杂查询
    # 新方案: 复用WHERE条件,COUNT和SELECT分别执行,避免不必要的子查询
    conditions = [OperatorAccount.deleted_at.is_(None)]

    # 搜索条件
    if search:
        search_filter = or_(
            OperatorAccount.username.ilike(f"%{search}%"),
            OperatorAccount.full_name.ilike(f"%{search}%"),
            OperatorAccount.email.ilike(f"%{search}%"),
            OperatorAccount.phone.ilike(f"%{search}%"),
        )
        conditions.append(search_filter)

    # 状态筛选
    if status_filter == "active":
        conditions.extend([
            OperatorAccount.is_active == True,
            OperatorAccount.is_locked == False
        ])
    elif status_filter == "inactive":
        conditions.append(OperatorAccount.is_active == False)
    elif status_filter == "locked":
        conditions.append(OperatorAccount.is_locked == True)

    # 计算总数 - 使用相同条件直接COUNT
    count_query = select(func.count(OperatorAccount.id)).where(*conditions)
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # 构建数据查询 - 使用相同条件
    query = select(OperatorAccount).where(*conditions)

    # 分页查询
    query = query.order_by(OperatorAccount.created_at.desc())
    query = query.limit(page_size).offset((page - 1) * page_size)

    result = await db.execute(query)
    operators = result.scalars().all()

    # 构建响应
    items = []
    for op in operators:
        items.append({
            "id": f"op_{op.id}",  # 添加前缀保持一致性
            "username": op.username,
            "full_name": op.full_name,
            "email": op.email,
            "phone": op.phone,
            "balance": float(op.balance),
            "customer_tier": op.customer_tier,
            "is_active": op.is_active,
            "is_locked": op.is_locked,
            "last_login_at": op.last_login_at.isoformat() if op.last_login_at else None,
            "created_at": op.created_at.isoformat() if op.created_at else None,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ==================== 手动充值 ====================


@router.post(
    "/recharge",
    response_model=RechargeResponse,
    status_code=status.HTTP_200_OK,
    responses={
        400: {"description": "请求参数错误或运营商不存在"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "运营商不存在"},
    },
    summary="手动充值",
    description="""
    财务人员为运营商手动充值。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **请求体**:
    - operator_id: 运营商ID
    - amount: 充值金额（必须大于0）
    - description: 充值备注（可选）
    - payment_proof: 付款凭证文件（可选）

    **响应数据**:
    - transaction_id: 交易记录ID
    - operator_id: 运营商ID
    - operator_name: 运营商名称
    - amount: 充值金额
    - balance_before: 充值前余额
    - balance_after: 充值后余额
    - description: 充值备注
    - created_at: 充值时间

    **业务规则**:
    - 充值金额必须大于0
    - 更新运营商余额
    - 创建充值类型交易记录
    - 记录财务操作审计日志
    - 支持上传付款凭证文件
    """,
)
async def manual_recharge(
    operator_id: str = Form(...),
    amount: float = Form(..., gt=0),
    description: Optional[str] = Form(None),
    payment_proof: Optional[UploadFile] = File(None),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> RechargeResponse:
    """手动充值API

    Args:
        operator_id: 运营商ID
        amount: 充值金额
        description: 充值备注
        payment_proof: 付款凭证文件
        token: JWT Token payload
        db: 数据库会话

    Returns:
        RechargeResponse: 充值结果

    Raises:
        HTTPException 400: 请求参数错误
        HTTPException 401: 未认证
        HTTPException 403: 权限不足
        HTTPException 404: 运营商不存在
        HTTPException 500: 服务器内部错误
    """
    # 从token中提取finance_id
    finance_id_str = token.get("sub")
    if not finance_id_str:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={"error_code": "INVALID_TOKEN", "message": "Token中缺少财务ID"},
        )

    try:
        from uuid import UUID
        finance_id = UUID(finance_id_str)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error_code": "INVALID_FINANCE_ID",
                "message": f"无效的财务ID格式: {finance_id_str}",
            },
        )

    try:
        from datetime import datetime
        from decimal import Decimal
        from sqlalchemy import select
        from ...models.operator import OperatorAccount
        from ...models.transaction import TransactionRecord
        from ...models.finance import FinanceOperationLog

        # 解析运营商ID（移除 "op_" 前缀）
        actual_operator_id = operator_id.replace("op_", "") if operator_id.startswith("op_") else operator_id

        try:
            operator_uuid = UUID(actual_operator_id)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "INVALID_OPERATOR_ID",
                    "message": f"无效的运营商ID格式: {operator_id}",
                },
            )

        # 查询运营商（使用行级锁防止并发问题）
        query = select(OperatorAccount).where(
            OperatorAccount.id == operator_uuid,
            OperatorAccount.deleted_at.is_(None)
        ).with_for_update()

        result = await db.execute(query)
        operator = result.scalar_one_or_none()

        if not operator:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error_code": "OPERATOR_NOT_FOUND", "message": "运营商不存在或已删除"},
            )

        # 验证运营商状态
        if not operator.is_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "OPERATOR_INACTIVE", "message": "运营商账户已停用"},
            )

        if operator.is_locked:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "OPERATOR_LOCKED", "message": "运营商账户已锁定"},
            )

        # 记录充值前余额
        balance_before = operator.balance
        recharge_amount = Decimal(str(amount))

        # 更新运营商余额
        operator.balance = operator.balance + recharge_amount
        balance_after = operator.balance

        # 处理付款凭证文件（如果有）
        payment_proof_path = None
        if payment_proof:
            import os
            upload_dir = "uploads/payment_proofs"
            os.makedirs(upload_dir, exist_ok=True)
            file_extension = payment_proof.filename.split('.')[-1] if payment_proof.filename else 'jpg'
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{actual_operator_id}.{file_extension}"
            file_path = os.path.join(upload_dir, filename)

            # 🚀 性能优化: 使用异步文件I/O,避免阻塞事件循环
            # 原方案: 同步写入文件,阻塞整个事件循环
            # 新方案: 异步写入文件,不阻塞其他请求
            content = await payment_proof.read()
            import aiofiles
            async with aiofiles.open(file_path, "wb") as f:
                await f.write(content)

            payment_proof_path = file_path

        # 创建交易记录（手动充值不需要设置payment_channel等字段）
        transaction = TransactionRecord(
            operator_id=operator_uuid,
            transaction_type="recharge",
            amount=recharge_amount,
            balance_before=balance_before,
            balance_after=balance_after,
            description=description or "财务手动充值",
        )
        db.add(transaction)

        # 记录财务操作日志
        operation_log = FinanceOperationLog(
            finance_account_id=finance_id,
            operation_type="manual_recharge",
            target_resource_type="operator",
            target_resource_id=operator_uuid,
            operation_details={
                "operator_id": str(operator_uuid),
                "operator_username": operator.username,
                "operator_name": operator.full_name,
                "amount": str(recharge_amount),
                "balance_before": str(balance_before),
                "balance_after": str(balance_after),
                "description": description,
                "payment_proof": payment_proof_path,
            },
            ip_address="127.0.0.1",  # TODO: 从请求中获取真实IP
        )
        db.add(operation_log)

        # 提交事务
        await db.commit()
        await db.refresh(transaction)

        # 返回充值结果
        return RechargeResponse(
            transaction_id=f"txn_{transaction.id}",
            operator_id=f"op_{operator_uuid}",
            operator_name=operator.full_name,
            amount=f"{recharge_amount:.2f}",
            balance_before=f"{balance_before:.2f}",
            balance_after=f"{balance_after:.2f}",
            description=description or "财务手动充值",
            created_at=transaction.created_at,
        )

    except BadRequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error_code": "RECHARGE_FAILED", "message": str(e)},
        )

    except NotFoundException as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error_code": "OPERATOR_NOT_FOUND", "message": str(e)},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"充值失败: {str(e)}",
            },
        )


@router.get(
    "/recharge-records",
    response_model=dict,
    status_code=status.HTTP_200_OK,
    summary="获取充值记录列表",
    description="财务人员查询所有运营商的充值记录，支持筛选和分页",
)
async def get_recharge_records(
    operator_id: Optional[str] = Query(None, description="运营商ID筛选（格式: op_xxx或uuid）"),
    start_date: Optional[str] = Query(None, description="开始日期（格式: YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（格式: YYYY-MM-DD）"),
    recharge_method: Optional[str] = Query(None, description="充值方式筛选(manual/online/bank_transfer)"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """获取充值记录列表

    返回所有运营商的充值记录，包括：
    - 手动充值（财务操作）
    - 在线充值（支付宝/微信支付）

    支持按运营商、日期范围筛选，按时间倒序排列
    """
    from datetime import datetime
    from sqlalchemy import select, and_, func
    from sqlalchemy.orm import selectinload
    from ...models.transaction import TransactionRecord
    from ...models.operator import OperatorAccount

    try:
        # 构建基础查询 - 查询充值和扣费类型的交易
        query = select(TransactionRecord).where(
            TransactionRecord.transaction_type.in_(["recharge", "deduct"])
        ).options(
            selectinload(TransactionRecord.operator)
        )

        # 筛选条件
        filters = []

        # 按运营商筛选
        if operator_id:
            # 解析运营商ID（移除op_前缀）
            actual_operator_id = operator_id.replace("op_", "") if operator_id.startswith("op_") else operator_id
            try:
                operator_uuid = UUID(actual_operator_id)
                filters.append(TransactionRecord.operator_id == operator_uuid)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_OPERATOR_ID", "message": "无效的运营商ID格式"}
                )

        # 按日期范围筛选
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                filters.append(TransactionRecord.created_at >= start_dt)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_DATE", "message": "开始日期格式错误，应为YYYY-MM-DD"}
                )

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                # 结束日期包含当天23:59:59
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
                filters.append(TransactionRecord.created_at <= end_dt)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_DATE", "message": "结束日期格式错误，应为YYYY-MM-DD"}
                )

        # 按充值方式筛选
        if recharge_method:
            if recharge_method == "manual":
                # 手动充值: payment_channel为空且transaction_type为recharge
                filters.append(and_(
                    TransactionRecord.payment_channel.is_(None),
                    TransactionRecord.transaction_type == 'recharge'
                ))
            elif recharge_method == "deduct":
                # 财务扣费: transaction_type为deduct
                filters.append(TransactionRecord.transaction_type == 'deduct')
            elif recharge_method == "online":
                # 在线充值: payment_channel不为空且不是bank_transfer
                filters.append(and_(
                    TransactionRecord.payment_channel.isnot(None),
                    TransactionRecord.payment_channel != 'bank_transfer'
                ))
            elif recharge_method == "bank_transfer":
                # 银行转账
                filters.append(TransactionRecord.payment_channel == 'bank_transfer')

        # 应用筛选条件
        if filters:
            query = query.where(and_(*filters))

        # 按时间倒序排列
        query = query.order_by(TransactionRecord.created_at.desc())

        # 统计总数
        count_query = select(func.count()).select_from(TransactionRecord).where(
            TransactionRecord.transaction_type.in_(["recharge", "deduct"])
        )
        if filters:
            count_query = count_query.where(and_(*filters))

        result = await db.execute(count_query)
        total = result.scalar()

        # 分页
        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        # 执行查询
        result = await db.execute(query)
        records = result.scalars().all()

        # 构建响应数据
        items = []
        for record in records:
            # 判断充值方式/交易类型
            if record.transaction_type == 'deduct':
                # 扣费类型
                recharge_method = "财务扣费"
                payment_info = None
            elif record.payment_channel == 'bank_transfer':
                recharge_method = "银行转账"
                payment_info = {
                    "channel": record.payment_channel,
                    "order_no": record.payment_order_no,
                    "status": record.payment_status,
                }
            elif record.payment_channel:
                recharge_method = "在线充值"
                payment_info = {
                    "channel": record.payment_channel,
                    "order_no": record.payment_order_no,
                    "status": record.payment_status,
                }
            else:
                recharge_method = "财务充值"
                payment_info = None

            # Generate human-readable transaction ID (format: TXN_YYYYMMDD_XXXXX)
            trans_created_time = record.created_at
            transaction_id = f"TXN_{trans_created_time.strftime('%Y%m%d')}_{str(record.id)[:5].upper()}"

            items.append({
                "id": transaction_id,
                "transaction_id": transaction_id,
                "operator_id": f"op_{record.operator_id}",
                "operator_username": record.operator.username,
                "operator_name": record.operator.full_name,
                "amount": f"{record.amount:.2f}",
                "balance_before": f"{record.balance_before:.2f}",
                "balance_after": f"{record.balance_after:.2f}",
                "recharge_method": recharge_method,
                "payment_info": payment_info,
                "description": record.description or "",
                "created_at": record.created_at.isoformat(),
            })

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
        }

    except HTTPException:
        raise

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "INTERNAL_ERROR",
                "message": f"查询充值记录失败: {str(e)}",
            },
        )


# ==================== 交易记录 ====================

@router.get(
    "/transactions",
    response_model=dict,
    status_code=status.HTTP_200_OK,
    responses={
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
    },
    summary="获取交易记录列表",
    description="""
    获取所有运营商的交易记录。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **支持筛选**:
    - operator_id: 运营商ID
    - transaction_type: 交易类型(recharge/consumption/refund)
    - start_time/end_time: 时间范围
    """
)
async def get_transactions(
    operator_id: Optional[str] = Query(None, description="运营商ID筛选"),
    transaction_type: Optional[str] = Query(None, description="交易类型筛选"),
    start_time: Optional[str] = Query(None, description="开始时间"),
    end_time: Optional[str] = Query(None, description="结束时间"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """获取交易记录列表"""
    from datetime import datetime
    from sqlalchemy import select, and_, or_, func
    from ...models import TransactionRecord, OperatorAccount

    try:
        # 构建查询条件 - 包含所有交易类型（充值、消费、退款、扣费）
        filters = [TransactionRecord.transaction_type.in_(['recharge', 'consumption', 'refund', 'deduct'])]

        # 按运营商筛选
        if operator_id:
            # 处理op_前缀或UUID格式
            if operator_id.startswith("op_"):
                operator_uuid = operator_id[3:]
            else:
                operator_uuid = operator_id

            try:
                operator_uuid = UUID(operator_uuid)
                filters.append(TransactionRecord.operator_id == operator_uuid)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_OPERATOR_ID", "message": "运营商ID格式错误"}
                )

        # 按交易类型筛选
        if transaction_type:
            if transaction_type not in ['recharge', 'consumption', 'refund', 'deduct']:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_TRANSACTION_TYPE", "message": "交易类型无效"}
                )
            filters.append(TransactionRecord.transaction_type == transaction_type)

        # 按时间范围筛选
        if start_time:
            try:
                start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                filters.append(TransactionRecord.created_at >= start_dt)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_START_TIME", "message": "开始时间格式错误"}
                )

        if end_time:
            try:
                end_dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
                filters.append(TransactionRecord.created_at <= end_dt)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"error_code": "INVALID_END_TIME", "message": "结束时间格式错误"}
                )

        # 查询总数
        count_stmt = select(func.count()).select_from(TransactionRecord).where(and_(*filters))
        total_result = await db.execute(count_stmt)
        total = total_result.scalar()

        # 查询分页数据
        stmt = (
            select(TransactionRecord, OperatorAccount)
            .join(OperatorAccount, TransactionRecord.operator_id == OperatorAccount.id)
            .where(and_(*filters))
            .order_by(TransactionRecord.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )

        result = await db.execute(stmt)
        records = result.all()

        # 构建返回数据
        items = []
        for record, operator in records:
            # Generate human-readable transaction ID
            trans_created_time = record.created_at
            transaction_id = f"TXN_{trans_created_time.strftime('%Y%m%d')}_{str(record.id)[:5].upper()}"

            items.append({
                "transaction_id": transaction_id,
                "operator_id": f"op_{record.operator_id}",
                "operator_name": operator.full_name,
                "transaction_type": record.transaction_type,
                "amount": f"{record.amount:.2f}",
                "balance_before": f"{record.balance_before:.2f}",
                "balance_after": f"{record.balance_after:.2f}",
                "description": record.description or "",
                "created_at": record.created_at.isoformat(),
            })

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error_code": "INTERNAL_ERROR", "message": f"查询交易记录失败: {str(e)}"}
        )


# ==================== 财务扣费 ====================

@router.post(
    "/deduct",
    response_model=dict,
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {"description": "请求参数错误"},
        401: {"description": "未认证或Token无效/过期"},
        403: {"description": "权限不足(非财务人员)"},
        404: {"description": "运营商不存在"},
    },
    summary="财务扣费",
    description="""
    财务人员为运营商扣费（用于纠正充值错误等情况）。

    **认证要求**:
    - Authorization: Bearer {JWT_TOKEN}
    - 用户类型: finance

    **请求参数**:
    - operator_id: 运营商ID
    - amount: 扣费金额（必须大于0且不超过运营商余额）
    - description: 扣费原因（必填）
    """
)
async def deduct_balance(
    request: Request,
    operator_id: str = Form(..., description="运营商ID"),
    amount: float = Form(..., description="扣费金额"),
    description: str = Form(..., description="扣费原因"),
    token: dict = Depends(require_finance),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """财务扣费API"""
    from decimal import Decimal
    from sqlalchemy import select
    from ...models import OperatorAccount, TransactionRecord, FinanceOperationLog

    try:
        # 验证金额
        if amount <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "INVALID_AMOUNT", "message": "扣费金额必须大于0"}
            )

        if not description or len(description.strip()) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "INVALID_DESCRIPTION", "message": "扣费原因不能为空"}
            )

        # 处理运营商ID
        if operator_id.startswith("op_"):
            operator_uuid = operator_id[3:]
        else:
            operator_uuid = operator_id

        try:
            operator_uuid = UUID(operator_uuid)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error_code": "INVALID_OPERATOR_ID", "message": "运营商ID格式错误"}
            )

        # 查询运营商
        stmt = select(OperatorAccount).where(OperatorAccount.id == operator_uuid)
        result = await db.execute(stmt)
        operator = result.scalar_one_or_none()

        if not operator:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error_code": "OPERATOR_NOT_FOUND", "message": "运营商不存在"}
            )

        # 检查余额是否足够
        deduct_amount = Decimal(str(amount))
        if operator.balance < deduct_amount:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "INSUFFICIENT_BALANCE",
                    "message": f"余额不足，当前余额: ¥{operator.balance:.2f}"
                }
            )

        # 记录扣费前后余额
        balance_before = operator.balance
        balance_after = balance_before - deduct_amount

        # 更新运营商余额
        operator.balance = balance_after

        # 创建交易记录（扣费使用deduct类型）
        transaction = TransactionRecord(
            operator_id=operator_uuid,
            transaction_type="deduct",  # 使用deduct类型表示财务扣费
            amount=deduct_amount,  # amount必须为正数,通过transaction_type区分增减
            balance_before=balance_before,
            balance_after=balance_after,
            description=f"财务扣费: {description}",
        )
        db.add(transaction)

        # 记录财务操作日志
        client_ip = request.client.host if request.client else "unknown"
        user_agent = request.headers.get("user-agent", "")

        operation_log = FinanceOperationLog(
            finance_account_id=UUID(token["sub"]),
            operation_type="deduct",
            target_resource_type="operator",
            target_resource_id=operator_uuid,
            operation_details={
                "amount": str(deduct_amount),
                "description": description,
                "balance_before": str(balance_before),
                "balance_after": str(balance_after),
            },
            ip_address=client_ip,
            user_agent=user_agent,
        )
        db.add(operation_log)

        await db.commit()
        await db.refresh(transaction)

        return {
            "transaction_id": f"TXN_{transaction.created_at.strftime('%Y%m%d')}_{str(transaction.id)[:5].upper()}",
            "operator_id": f"op_{operator_uuid}",
            "operator_name": operator.full_name,
            "amount": f"{deduct_amount:.2f}",
            "balance_before": f"{balance_before:.2f}",
            "balance_after": f"{balance_after:.2f}",
            "description": f"财务扣费: {description}",
            "created_at": transaction.created_at.isoformat(),
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error_code": "INTERNAL_ERROR", "message": f"扣费失败: {str(e)}"}
        )


# ==================== 银行转账审核 (T189f-T189h) ====================

from .bank_transfers import router as bank_transfers_router

# Include bank transfers router
router.include_router(bank_transfers_router)
